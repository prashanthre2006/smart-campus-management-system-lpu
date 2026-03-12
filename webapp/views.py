import base64
import json
import os
import re
from datetime import datetime, timedelta

import cv2
import numpy as np
from django.contrib import auth
from django.forms import modelform_factory
from django.db.models.deletion import ProtectedError
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from django.db import models, IntegrityError
from django.utils import timezone

from config import (
    ATTENDANCE_CSV,
    BLUR_THRESHOLD,
    BRIGHTNESS_MAX,
    BRIGHTNESS_MIN,
    CONFIDENCE_THRESHOLD,
    CONFIDENCE_STRICT,
    RECOGNITION_ACCEPT_MAX,
    DATA_DIR,
    MODEL_DIR,
    REQUIRED_HITS,
    PERIODS,
    ADMIN_PASSWORD,
    ADMIN_USERNAME,
    ALERTS_ENABLED,
    ALERTS_ON_MARKED,
    ALERTS_ON_UNKNOWN,
    PHONEPE_WEBHOOK_TOKEN,
    QR_SCANNER_TOKEN,
    REQUIRED_ENROLL_IMAGES,
)
from db import init_db
from utils import brightness_level, ensure_dir, enhance_low_light, is_blurry, load_labels, mark_attendance
from .models import Attendance, AuditLog, Device, Student, UserProfile, SchedulePeriod
from .models import (
    BreakTimeSlot,
    CampusBlock,
    Classroom,
    ClassSection,
    Course,
    Faculty,
    FoodItem,
    FoodOrder,
    FoodOrderItem,
    MakeupAttendance,
    MakeupClassSession,
    Payment,
    Subject,
    SubjectOffering,
    TimetableEntry,
)


def _user_is_admin(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    profile = UserProfile.objects.filter(user=user).first()
    return profile and profile.role == "admin"


def _user_is_faculty(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    profile = UserProfile.objects.filter(user=user).first()
    return bool(profile and profile.role in {"faculty", "admin"})


def _has_admin_access(request):
    return bool(request.session.get("is_admin")) or _user_is_admin(request.user)


def _has_faculty_access(request):
    return bool(request.session.get("lecturer_user_id")) or _has_admin_access(request) or _user_is_faculty(request.user)


def _mask_username(username: str) -> str:
    if not username:
        return "********"
    if len(username) <= 2:
        return "*" * len(username)
    return ("*" * (len(username) - 2)) + username[-2:]


MODEL_PATH = os.path.join(MODEL_DIR, "face_model.yml")
LABELS_PATH = os.path.join(MODEL_DIR, "labels.json")

_last_hits = {}
_last_marked = {}
_last_status = "Ready"
_rate_limit_buckets = {}


def _reset_hits_except(name_to_keep=None):
    for n in list(_last_hits.keys()):
        if name_to_keep is None or n != name_to_keep:
            _last_hits[n] = 0


def _normalize_face(gray_face):
    resized = cv2.resize(gray_face, (200, 200))
    return cv2.equalizeHist(resized)


def _rate_limited(key: str, limit: int, window_sec: int) -> bool:
    now_ts = datetime.now().timestamp()
    bucket = _rate_limit_buckets.get(key, [])
    bucket = [ts for ts in bucket if (now_ts - ts) <= window_sec]
    if len(bucket) >= limit:
        _rate_limit_buckets[key] = bucket
        return True
    bucket.append(now_ts)
    _rate_limit_buckets[key] = bucket
    return False


def _send_notification(subject: str, body: str):
    if ALERTS_ENABLED:
        AuditLog.objects.create(actor="system", action="alert_event", detail=f"{subject}: {body[:180]}")


def _log_unknown_face(client_ip: str, reason: str):
    # Avoid flooding audit logs when camera is running continuously.
    key = f"unknown_face:{client_ip}:{reason}"
    if _rate_limited(key, limit=1, window_sec=10):
        return
    AuditLog.objects.create(actor="system", action="unknown_face", detail=f"{reason} from {client_ip}")


def _suggest_next_slots(day: str, period, room: str, faculty, section, limit: int = 3):
    candidates = SchedulePeriod.objects.order_by("start_time")
    suggestions = []
    seen_labels = set()
    for cand in candidates:
        if not cand.start_time or not cand.end_time or cand.end_time <= cand.start_time:
            continue
        start_mins = (cand.start_time.hour * 60) + cand.start_time.minute
        end_mins = (cand.end_time.hour * 60) + cand.end_time.minute
        # Keep only exact 1-hour slots aligned to hour boundaries.
        if cand.start_time.minute != 0 or cand.end_time.minute != 0 or (end_mins - start_mins) != 60:
            continue
        conflicts = TimetableEntry.objects.select_related("offering__faculty", "offering__class_section").filter(
            day=day, period=cand
        )
        if conflicts.filter(room=room).exists():
            continue
        if faculty and conflicts.filter(offering__faculty=faculty).exists():
            continue
        if section and conflicts.filter(offering__class_section=section).exists():
            continue
        slot_label = f"{cand.start_time.hour}-{cand.end_time.hour}"
        if slot_label in seen_labels:
            continue
        seen_labels.add(slot_label)
        suggestions.append(slot_label)
        if len(suggestions) >= limit:
            break
    return suggestions


def welcome(request):
    return render(request, "welcome.html")


def _ensure_default_food_data():
    if not BreakTimeSlot.objects.exists():
        default_slots = [
            ("Morning Break", "10:00", "10:30"),
            ("Lunch Break", "13:00", "14:00"),
            ("Evening Break", "16:30", "17:00"),
        ]
        for name, start, end in default_slots:
            BreakTimeSlot.objects.create(
                name=name,
                start_time=datetime.strptime(start, "%H:%M").time(),
                end_time=datetime.strptime(end, "%H:%M").time(),
                is_active=True,
            )
    if not FoodItem.objects.exists():
        default_items = [
            ("Idli Sambar", "Breakfast", "Steamed idlis served with sambar and chutney.", 45, 8),
            ("Aloo Paratha", "Breakfast", "Fresh paratha with curd and pickle.", 60, 10),
            ("Veg Thali", "Main Course", "Balanced meal with rice, chapati, dal, and curry.", 110, 15),
            ("Paneer Butter Masala", "Main Course", "Creamy paneer gravy served with roti.", 140, 16),
            ("Veg Burger", "Snacks", "Crispy veg patty burger with fries.", 90, 12),
            ("French Fries", "Snacks", "Crisp salted fries.", 70, 8),
            ("Cold Coffee", "Beverages", "Chilled coffee with milk and ice cream.", 80, 6),
            ("Fresh Lime Soda", "Beverages", "Sweet and salted lime drink.", 50, 5),
            ("Gulab Jamun", "Desserts", "Warm gulab jamun (2 pcs).", 55, 5),
            ("Brownie", "Desserts", "Chocolate brownie slice.", 65, 5),
        ]
        for name, category, description, price, prep in default_items:
            FoodItem.objects.create(
                name=name,
                category=category,
                description=description,
                price=price,
                prep_time_mins=prep,
                is_available=True,
            )


def _ensure_default_period_slots():
    standard_ranges = [(9, 10), (10, 11), (11, 12), (12, 13), (13, 14), (14, 15), (15, 16), (16, 17)]
    for start_hr, end_hr in standard_ranges:
        start_time = datetime.strptime(f"{start_hr:02d}:00", "%H:%M").time()
        end_time = datetime.strptime(f"{end_hr:02d}:00", "%H:%M").time()
        period = SchedulePeriod.objects.filter(start_time=start_time, end_time=end_time).first()
        if not period:
            SchedulePeriod.objects.create(
                name=f"{start_hr}-{end_hr}",
                start_time=start_time,
                end_time=end_time,
                days_of_week="Mon,Tue,Wed,Thu,Fri,Sat",
            )


def student(request):
    init_db("data/attendance.db")
    today_day = datetime.now().strftime("%a")
    timetable_entries = (
        TimetableEntry.objects.select_related(
            "period", "offering__subject", "offering__faculty", "offering__class_section"
        )
        .filter(day=today_day)
        .order_by("period__start_time")
    )
    today = datetime.now().date()
    makeup_sessions = MakeupClassSession.objects.select_related("faculty", "course", "class_section").filter(
        date__gte=today
    ).order_by("date", "start_time")[:30]
    return render(
        request,
        "student.html",
        {
            "timetable_entries": timetable_entries,
            "today_day": today_day,
            "makeup_sessions": makeup_sessions,
        },
    )


def admin_page(request):
    init_db("data/attendance.db")
    return render(
        request,
        "admin.html",
        {
            "periods": PERIODS,
            "is_admin": bool(request.session.get("is_admin")),
            "masked_admin_username": _mask_username(ADMIN_USERNAME),
        },
    )


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "")
        password = request.POST.get("password", "")
        user = auth.authenticate(request, username=username, password=password)
        if user:
            auth.login(request, user)
            if _user_is_admin(user):
                return redirect("dashboard")
            if _user_is_faculty(user):
                return redirect("makeup_class_page")
            return redirect("student")
        return render(request, "login.html", {"error": "Invalid credentials"})
    return render(request, "login.html")


def logout_view(request):
    auth.logout(request)
    return redirect("login")


def dashboard(request):
    if not _has_admin_access(request):
        return redirect("login")
    return render(request, "dashboard.html")


def devices_page(request):
    if not _has_admin_access(request):
        return redirect("login")
    return render(request, "devices.html")


def schedule_page(request):
    if not _has_admin_access(request):
        return redirect("login")
    _ensure_default_period_slots()
    blocks = CampusBlock.objects.all().order_by("code", "name")
    periods = (
        SchedulePeriod.objects.filter(
            start_time__in=[
                datetime.strptime("09:00", "%H:%M").time(),
                datetime.strptime("10:00", "%H:%M").time(),
                datetime.strptime("11:00", "%H:%M").time(),
                datetime.strptime("12:00", "%H:%M").time(),
                datetime.strptime("13:00", "%H:%M").time(),
                datetime.strptime("14:00", "%H:%M").time(),
                datetime.strptime("15:00", "%H:%M").time(),
                datetime.strptime("16:00", "%H:%M").time(),
            ]
        )
        .order_by("start_time")
    )
    status_msg = ""
    error_msg = ""

    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action == "delete_timetable":
            row_id = request.POST.get("id", "").strip()
            if row_id:
                TimetableEntry.objects.filter(id=row_id).delete()
                status_msg = "Timetable row deleted."
        else:
            day = request.POST.get("day", "").strip()
            faculty_name = request.POST.get("faculty_name", "").strip()
            course_name = request.POST.get("course_name", "").strip()
            course_code = request.POST.get("course_code", "").strip()
            block_id = request.POST.get("block_id", "").strip()
            period_id = request.POST.get("period_id", "").strip()
            classroom_no = request.POST.get("classroom_no", "").strip()

            if not day or not faculty_name or not course_name or not course_code or not block_id or not period_id or not classroom_no:
                error_msg = "Faculty name, course name, course code, block, classroom no, day, and slot are required."
            else:
                block = CampusBlock.objects.filter(id=block_id).first()
                period = SchedulePeriod.objects.filter(id=period_id).first()
                if not block or not period:
                    error_msg = "Invalid block or slot selected."
                else:
                    faculty = Faculty.objects.filter(name=faculty_name).first() or Faculty.objects.create(
                        name=faculty_name
                    )
                    course = Course.objects.filter(code=course_code).first()
                    if not course:
                        course = Course.objects.create(name=course_name, code=course_code)
                    subject = Subject.objects.filter(code=course_code).first()
                    if not subject:
                        subject = Subject.objects.create(name=course_name, code=course_code)
                    section_name = f"{block.code}-{classroom_no}"
                    section = ClassSection.objects.filter(name=section_name).first()
                    if not section:
                        section = ClassSection.objects.create(name=section_name, program="Campus Block")
                    offering = SubjectOffering.objects.filter(
                        course=course, class_section=section, subject=subject
                    ).first()
                    if not offering:
                        offering = SubjectOffering.objects.create(
                            course=course,
                            class_section=section,
                            subject=subject,
                            faculty=faculty,
                        )
                    else:
                        offering.faculty = faculty
                        offering.save()

                    conflict_qs = TimetableEntry.objects.select_related("offering__faculty", "offering__class_section").filter(
                        day=day,
                        period=period,
                    )
                    room_conflict = conflict_qs.filter(room=f"{block.code}-{classroom_no}").first()
                    faculty_conflict = conflict_qs.filter(offering__faculty=faculty).first()
                    section_conflict = conflict_qs.filter(offering__class_section=section).first()
                    suggested_slots = _suggest_next_slots(
                        day=day,
                        period=period,
                        room=f"{block.code}-{classroom_no}",
                        faculty=faculty,
                        section=section,
                        limit=3,
                    )
                    suggestion_hint = f" Suggested slots: {', '.join(suggested_slots)}." if suggested_slots else ""

                    if room_conflict:
                        error_msg = (
                            f"Room conflict: {block.code}-{classroom_no} is already assigned for {day} "
                            f"slot {period.start_time.hour}-{period.end_time.hour}.{suggestion_hint}"
                        )
                    elif faculty_conflict:
                        error_msg = (
                            f"Faculty conflict: {faculty.name} already has a class on {day} "
                            f"slot {period.start_time.hour}-{period.end_time.hour}.{suggestion_hint}"
                        )
                    elif section_conflict:
                        error_msg = (
                            f"Section conflict: {section.name} already has a class on {day} "
                            f"slot {period.start_time.hour}-{period.end_time.hour}.{suggestion_hint}"
                        )
                    else:
                        try:
                            TimetableEntry.objects.create(
                                day=day,
                                period=period,
                                offering=offering,
                                room=f"{block.code}-{classroom_no}",
                            )
                            status_msg = "Timetable added successfully."
                        except IntegrityError:
                            error_msg = "Same timetable entry already exists for this day and slot."

    entries = TimetableEntry.objects.select_related(
        "period", "offering__course", "offering__subject", "offering__class_section", "offering__faculty"
    ).order_by("day", "period__start_time")
    day_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    grouped = {d: [] for d in day_order}
    for e in entries:
        grouped[e.day].append(e)
    grouped_entries = [{"day": d, "entries": grouped[d]} for d in day_order]
    return render(
        request,
        "schedule.html",
        {
            "blocks": blocks,
            "periods": periods,
            "slot_options": [
                {"id": p.id, "label": f"{p.start_time.hour}-{p.end_time.hour}"} for p in periods
            ],
            "grouped_entries": grouped_entries,
            "days": day_order,
            "status_msg": status_msg,
            "error_msg": error_msg,
        },
    )


def defaulters_page(request):
    if not _has_admin_access(request):
        return redirect("login")
    return render(request, "defaulters.html")


def monthly_page(request):
    if not _has_admin_access(request):
        return redirect("login")
    return render(request, "monthly.html")


MANAGE_ENTITY_CONFIG = {
    "faculty": {"model": Faculty, "fields": ["name", "department"], "title": "Faculty"},
    "student": {"model": Student, "fields": ["name", "reg_no"], "title": "Students"},
    "subject": {"model": Subject, "fields": ["name", "code"], "title": "Subjects"},
    "course": {"model": Course, "fields": ["name", "code"], "title": "Courses"},
    "scheduleperiod": {
        "model": SchedulePeriod,
        "fields": ["name", "start_time", "end_time", "days_of_week"],
        "title": "Time Slots",
    },
    "timetableentry": {"model": TimetableEntry, "fields": ["day", "period", "offering", "room"], "title": "Timetable"},
    "device": {
        "model": Device,
        "fields": ["name", "location", "is_active", "kiosk_id", "status"],
        "title": "Devices",
    },
    "auditlog": {"model": AuditLog, "fields": ["actor", "action", "detail"], "title": "Audit Logs"},
    "campusblock": {"model": CampusBlock, "fields": ["name", "code", "capacity"], "title": "Campus Blocks"},
    "classroom": {
        "model": Classroom,
        "fields": ["block", "name", "code", "capacity", "is_active"],
        "title": "Classrooms",
    },
    "breaktimeslot": {
        "model": BreakTimeSlot,
        "fields": ["name", "start_time", "end_time", "is_active"],
        "title": "Break Slots",
    },
    "fooditem": {
        "model": FoodItem,
        "fields": ["name", "category", "description", "price", "is_available", "prep_time_mins"],
        "title": "Food Items",
    },
    "foodorder": {
        "model": FoodOrder,
        "fields": ["student", "slot", "order_mode", "delivery_location", "status", "total_amount"],
        "title": "Food Orders",
    },
    "makeupclasssession": {
        "model": MakeupClassSession,
        "fields": ["faculty", "course", "class_section", "date", "start_time", "end_time", "remedial_code", "notes"],
        "title": "Make-Up Sessions",
    },
    "makeupattendance": {
        "model": MakeupAttendance,
        "fields": ["session", "student", "mode"],
        "title": "Make-Up Attendance",
    },
}


def manage_entity_page(request, entity):
    if not _has_admin_access(request):
        return redirect("login")
    if entity == "payment":
        return redirect("payments_page")
    cfg = MANAGE_ENTITY_CONFIG.get(entity)
    if not cfg:
        return redirect("admin_page")
    model = cfg["model"]
    fields = cfg["fields"]
    FormClass = modelform_factory(model, fields=fields)
    form = FormClass()
    status_msg = ""
    error_msg = ""

    if request.method == "POST":
        action = request.POST.get("action", "add")
        if action == "delete":
            row_id = request.POST.get("id", "")
            if row_id:
                try:
                    model.objects.filter(id=row_id).delete()
                    status_msg = "Record deleted."
                except ProtectedError:
                    error_msg = "Cannot delete this record because it is used by other data."
                except Exception as ex:
                    error_msg = f"Delete failed: {str(ex)}"
        else:
            form = FormClass(request.POST)
            if form.is_valid():
                try:
                    form.save()
                    status_msg = "Record saved."
                    form = FormClass()
                except Exception as ex:
                    error_msg = f"Save failed: {str(ex)}"
            else:
                error_msg = "Please fix the form errors."

    records = model.objects.all().order_by("-id")[:100]
    columns = ["id"] + fields
    display_rows = []
    for rec in records:
        row_vals = []
        for col in columns:
            val = getattr(rec, col, "")
            row_vals.append(str(val) if val is not None else "-")
        display_rows.append({"id": rec.id, "values": row_vals})
    return render(
        request,
        "manage_entity.html",
        {
            "title": cfg["title"],
            "entity": entity,
            "form": form,
            "display_rows": display_rows,
            "columns": columns,
            "status_msg": status_msg,
            "error_msg": error_msg,
        },
    )


def smart_food_page(request):
    _ensure_default_food_data()
    orders = FoodOrder.objects.select_related("slot", "student").order_by("-ordered_at")[:8]
    slots = BreakTimeSlot.objects.filter(is_active=True).order_by("start_time")
    items = FoodItem.objects.filter(is_available=True).order_by("category", "name")
    menu_sections = {}
    for item in items:
        menu_sections.setdefault(item.category, []).append(item)
    peak_slots = (
        FoodOrder.objects.values("slot__name")
        .annotate(order_count=models.Count("id"))
        .order_by("-order_count")[:5]
    )
    payment_stats = {
        "total_payments": Payment.objects.count(),
        "paid_amount": Payment.objects.filter(status="paid").aggregate(total=models.Sum("amount"))["total"] or 0,
        "pending_payments": Payment.objects.filter(status="pending").count(),
        "failed_payments": Payment.objects.filter(status="failed").count(),
    }
    return render(
        request,
        "smart_food.html",
        {
            "orders": orders,
            "slots": slots,
            "menu_sections": menu_sections,
            "peak_slots": peak_slots,
            "payment_stats": payment_stats,
            "item_count": items.count(),
        },
    )


def smart_food_menu_page(request):
    _ensure_default_food_data()
    menu_sections = {}
    for item in FoodItem.objects.filter(is_available=True).order_by("category", "name"):
        menu_sections.setdefault(item.category, []).append(item)
    return render(
        request,
        "smart_food_menu.html",
        {"menu_sections": menu_sections},
    )


def smart_food_checkout_page(request):
    _ensure_default_food_data()
    slots = BreakTimeSlot.objects.filter(is_active=True).order_by("start_time")
    students = Student.objects.order_by("name")[:300]
    return render(
        request,
        "smart_food_checkout.html",
        {
            "slots": slots,
            "students": students,
        },
    )


def smart_food_orders_page(request):
    _ensure_default_food_data()
    orders = FoodOrder.objects.select_related("slot", "student").order_by("-ordered_at")[:40]
    peak_slots = (
        FoodOrder.objects.values("slot__name")
        .annotate(order_count=models.Count("id"))
        .order_by("-order_count")[:8]
    )
    return render(
        request,
        "smart_food_orders.html",
        {
            "orders": orders,
            "peak_slots": peak_slots,
        },
    )


def smart_food_kitchen_page(request):
    if not _has_admin_access(request):
        return redirect("login")
    orders = FoodOrder.objects.select_related("slot", "student").order_by("-ordered_at")[:80]
    status_counts = (
        FoodOrder.objects.values("status")
        .annotate(count=models.Count("id"))
        .order_by("status")
    )
    return render(
        request,
        "smart_food_kitchen.html",
        {
            "orders": orders,
            "status_counts": status_counts,
            "status_choices": ["pending", "confirmed", "preparing", "ready", "collected", "cancelled"],
        },
    )


def campus_resource_page(request):
    blocks = CampusBlock.objects.all().order_by("code")
    classrooms = Classroom.objects.select_related("block").all()
    total_classroom_capacity = classrooms.aggregate(total=models.Sum("capacity"))["total"] or 0
    total_block_capacity = blocks.aggregate(total=models.Sum("capacity"))["total"] or 0

    classroom_utilization = 0
    if total_classroom_capacity:
        classroom_utilization = round((Attendance.objects.count() / total_classroom_capacity) * 100, 2)

    faculty_workload = (
        Faculty.objects.values("name")
        .annotate(
            assigned_subjects=models.Count("subjectoffering", distinct=True),
            timetable_slots=models.Count("subjectoffering__timetableentry", distinct=True),
        )
        .order_by("-timetable_slots", "-assigned_subjects")
    )
    block_utilization = []
    for block in blocks:
        block_rooms = classrooms.filter(block=block)
        block_capacity = block_rooms.aggregate(total=models.Sum("capacity"))["total"] or 0
        utilization = 0
        if total_classroom_capacity:
            utilization = round((block_capacity / total_classroom_capacity) * 100, 2)
        block_utilization.append(
            {
                "code": block.code,
                "name": block.name,
                "capacity": block_capacity,
                "utilization": utilization,
            }
        )
    return render(
        request,
        "campus_resource.html",
        {
            "blocks": blocks,
            "classrooms": classrooms[:20],
            "block_utilization": block_utilization,
            "total_block_capacity": total_block_capacity,
            "total_classroom_capacity": total_classroom_capacity,
            "classroom_utilization": classroom_utilization,
            "faculty_workload": faculty_workload,
        },
    )


def makeup_class_page(request):
    sessions = MakeupClassSession.objects.select_related("faculty", "course", "class_section").all()[:20]
    attendance_summary = (
        MakeupAttendance.objects.values("session__remedial_code")
        .annotate(total_present=models.Count("id"))
        .order_by("-total_present")[:10]
    )
    return render(
        request,
        "makeup_class.html",
        {
            "sessions": sessions,
            "attendance_summary": attendance_summary,
            "total_sessions": MakeupClassSession.objects.count(),
            "total_makeup_attendance": MakeupAttendance.objects.count(),
            "lecturer_logged_in": bool(request.session.get("lecturer_user_id") or request.session.get("is_admin")),
            "lecturer_name": request.session.get("lecturer_name", ""),
        },
    )


def faculty_workbench_page(request):
    if not _has_faculty_access(request):
        return redirect("login")
    today = datetime.now().date()
    today_day = datetime.now().strftime("%a")
    faculty_name = request.session.get("lecturer_name", "") or (
        request.user.get_username() if request.user and request.user.is_authenticated else ""
    )
    faculty = None
    if faculty_name:
        faculty = Faculty.objects.filter(name__iexact=faculty_name).first()
    if not faculty and request.user and request.user.is_authenticated:
        faculty = Faculty.objects.filter(user=request.user).first()

    timetable_base_qs = TimetableEntry.objects.select_related(
        "period", "offering__subject", "offering__course", "offering__class_section", "offering__faculty"
    )
    timetable_qs = timetable_base_qs.filter(day=today_day)
    if faculty:
        timetable_qs = timetable_qs.filter(offering__faculty=faculty)

    makeup_qs = MakeupClassSession.objects.select_related("course", "class_section").filter(date__gte=today)
    if faculty:
        makeup_qs = makeup_qs.filter(faculty=faculty)

    # Suggest free slots for next make-up planning.
    free_slot_suggestions = []
    seen_day_slot_labels = set()
    raw_slots = list(SchedulePeriod.objects.order_by("start_time", "end_time", "id"))
    # Deduplicate by start/end and ignore invalid ranges like 13-13.
    seen_slot_ranges = set()
    all_slots = []
    for slot in raw_slots:
        if not slot.start_time or not slot.end_time:
            continue
        if slot.end_time <= slot.start_time:
            continue
        start_mins = (slot.start_time.hour * 60) + slot.start_time.minute
        end_mins = (slot.end_time.hour * 60) + slot.end_time.minute
        if slot.start_time.minute != 0 or slot.end_time.minute != 0 or (end_mins - start_mins) != 60:
            continue
        key = (slot.start_time, slot.end_time)
        if key in seen_slot_ranges:
            continue
        seen_slot_ranges.add(key)
        all_slots.append(slot)
    for offset in range(0, 5):
        dt = today + timedelta(days=offset)
        day_code = dt.strftime("%a")
        used_qs = timetable_base_qs.filter(day=day_code)
        if faculty:
            used_qs = used_qs.filter(offering__faculty=faculty)
        used_period_ids = set(used_qs.values_list("period_id", flat=True))
        for slot in all_slots:
            if slot.id not in used_period_ids:
                slot_label = f"{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}"
                day_slot_key = (dt.strftime("%Y-%m-%d"), slot_label)
                if day_slot_key in seen_day_slot_labels:
                    continue
                seen_day_slot_labels.add(day_slot_key)
                free_slot_suggestions.append(
                    {
                        "date": dt.strftime("%Y-%m-%d"),
                        "day": day_code,
                        "slot": slot_label,
                    }
                )
            if len(free_slot_suggestions) >= 8:
                break
        if len(free_slot_suggestions) >= 8:
            break

    return render(
        request,
        "faculty_workbench.html",
        {
            "faculty_name": faculty.name if faculty else (faculty_name or "Faculty"),
            "today_day": today_day,
            "today_timetable": timetable_qs.order_by("period__start_time")[:20],
            "upcoming_makeups": makeup_qs.order_by("date", "start_time")[:20],
            "free_slot_suggestions": free_slot_suggestions,
        },
    )


def student_profile_page(request):
    if not request.user.is_authenticated and not _has_admin_access(request):
        return redirect("login")
    students = Student.objects.order_by("name")[:500]
    selected_name = request.GET.get("student_name", "").strip()
    if request.user.is_authenticated and not (_user_is_admin(request.user) or _user_is_faculty(request.user)):
        own_student = Student.objects.filter(user=request.user).first()
        if own_student:
            selected_name = own_student.name
            students = Student.objects.filter(id=own_student.id)
    selected_student = Student.objects.filter(name=selected_name).first() if selected_name else None
    stats = None
    recent_records = []
    subject_rows = []

    if selected_student:
        all_records = Attendance.objects.filter(student=selected_student)
        total = all_records.count()
        present_count = all_records.filter(status="present").count()
        late_count = all_records.filter(status="late").count()
        absent_count = all_records.filter(status="absent").count()
        attendance_rate = round(((present_count + late_count) / total) * 100, 2) if total else 0
        makeup_marked = MakeupAttendance.objects.filter(student=selected_student).count()
        food_orders = FoodOrder.objects.filter(student=selected_student)
        food_order_count = food_orders.count()
        food_spend = food_orders.aggregate(total=models.Sum("total_amount"))["total"] or 0

        summary = {}
        for rec in all_records.order_by("-date", "-marked_at"):
            parts = [p.strip() for p in rec.period_name.split(" - ")]
            subject_key = parts[2] if len(parts) >= 3 else parts[-1]
            if subject_key not in summary:
                summary[subject_key] = {"subject": subject_key, "present": 0, "late": 0, "absent": 0, "total": 0}
            summary[subject_key][rec.status] = summary[subject_key].get(rec.status, 0) + 1
            summary[subject_key]["total"] += 1

        subject_rows = sorted(summary.values(), key=lambda r: r["subject"].lower())
        regular_recent = list(
            all_records.select_related("device").order_by("-marked_at")[:30]
        )
        makeup_recent = list(
            MakeupAttendance.objects.select_related("session")
            .filter(student=selected_student)
            .order_by("-marked_at")[:30]
        )

        recent_records = []
        for r in regular_recent:
            recent_records.append(
                {
                    "date": r.date,
                    "period_name": r.period_name,
                    "status": r.status,
                    "device_name": r.device.name if r.device else "Webcam",
                    "marked_at": r.marked_at,
                }
            )
        for m in makeup_recent:
            session = m.session
            recent_records.append(
                {
                    "date": session.date if session else None,
                    "period_name": f"Make-Up | {session.remedial_code}" if session else "Make-Up",
                    "status": "present",
                    "device_name": "Webcam",
                    "marked_at": m.marked_at,
                }
            )
        recent_records.sort(key=lambda x: x["marked_at"], reverse=True)
        recent_records = recent_records[:30]
        stats = {
            "total": total,
            "present": present_count,
            "late": late_count,
            "absent": absent_count,
            "attendance_rate": attendance_rate,
            "makeup_marked": makeup_marked,
            "food_order_count": food_order_count,
            "food_spend": food_spend,
        }

    return render(
        request,
        "student_profile.html",
        {
            "students": students,
            "selected_name": selected_name,
            "selected_student": selected_student,
            "stats": stats,
            "subject_rows": subject_rows,
            "recent_records": recent_records,
        },
    )


def payments_page(request):
    payments = Payment.objects.select_related("order", "order__student").all()[:30]
    paid_total = Payment.objects.filter(status="paid").aggregate(total=models.Sum("amount"))["total"] or 0
    pending_count = Payment.objects.filter(status="pending").count()
    failed_count = Payment.objects.filter(status="failed").count()
    thirty_mins_ago = timezone.now() - timedelta(minutes=30)
    aged_pending_count = Payment.objects.filter(status="pending", order__ordered_at__lte=thirty_mins_ago).count()
    return render(
        request,
        "payments.html",
        {
            "payments": payments,
            "paid_total": paid_total,
            "payment_count": Payment.objects.count(),
            "pending_count": pending_count,
            "failed_count": failed_count,
            "aged_pending_count": aged_pending_count,
            "is_admin": _has_admin_access(request),
        },
    )


def ops_center_page(request):
    if not _has_admin_access(request):
        return redirect("login")
    date_str = request.GET.get("date", "").strip()
    if date_str:
        try:
            selected_day = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            selected_day = datetime.now().date()
    else:
        selected_day = datetime.now().date()

    low_conf_logs = AuditLog.objects.filter(action="low_conf_mark", created_at__date=selected_day).order_by("-created_at")[:40]
    dispute_logs = AuditLog.objects.filter(action="payment_dispute", created_at__date=selected_day).order_by("-created_at")[:40]
    refund_logs = AuditLog.objects.filter(action="payment_refund", created_at__date=selected_day).order_by("-created_at")[:40]
    unknown_face_logs = AuditLog.objects.filter(action="unknown_face", created_at__date=selected_day).order_by("-created_at")[:40]
    pending_payment_reviews = Payment.objects.filter(status="pending").count()
    recent_audit = AuditLog.objects.filter(created_at__date=selected_day).order_by("-created_at")[:30]
    return render(
        request,
        "ops_center.html",
        {
            "low_conf_logs": low_conf_logs,
            "dispute_logs": dispute_logs,
            "refund_logs": refund_logs,
            "unknown_face_logs": unknown_face_logs,
            "pending_payment_reviews": pending_payment_reviews,
            "recent_audit": recent_audit,
            "selected_day": selected_day.strftime("%Y-%m-%d"),
        },
    )


def payment_success_page(request):
    order_id = request.GET.get("order_id", "")
    return render(request, "payment_success.html", {"order_id": order_id})


def _decode_image(data_url: str):
    if "," in data_url:
        data_url = data_url.split(",", 1)[1]
    img_data = base64.b64decode(data_url)
    nparr = np.frombuffer(img_data, np.uint8)
    return cv2.imdecode(nparr, cv2.IMREAD_COLOR)


@csrf_exempt
def api_enroll(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)

    name = request.POST.get("name", "").strip()
    if not name:
        return JsonResponse({"ok": False, "error": "Name required"}, status=400)

    data_url = request.POST.get("frame", "")
    if not data_url:
        return JsonResponse({"ok": False, "error": "Frame required"}, status=400)

    img = _decode_image(data_url)
    if img is None:
        return JsonResponse({"ok": False, "error": "Invalid image"}, status=400)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(
        os.path.join(cv2.data.haarcascades, "haarcascade_frontalface_default.xml")
    )
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.15, minNeighbors=5, minSize=(60, 60))
    if len(faces) == 0:
        enhanced = cv2.equalizeHist(gray)
        faces = face_cascade.detectMultiScale(enhanced, scaleFactor=1.1, minNeighbors=4, minSize=(45, 45))
    if len(faces) == 0:
        return JsonResponse({"ok": False, "error": "No face detected"})
    if len(faces) > 1:
        return JsonResponse({"ok": False, "error": "Multiple faces detected. Keep only one face in camera."})

    (x, y, w, h) = max(faces, key=lambda f: f[2] * f[3])
    face_roi = gray[y : y + h, x : x + w]

    bright = brightness_level(face_roi)
    if bright < BRIGHTNESS_MIN:
        face_roi = enhance_low_light(face_roi)
        bright = brightness_level(face_roi)
    blurry, _ = is_blurry(face_roi, BLUR_THRESHOLD)
    if blurry or bright < BRIGHTNESS_MIN or bright > BRIGHTNESS_MAX:
        return JsonResponse({"ok": False, "error": "Low quality frame"})
    face_roi = _normalize_face(face_roi)

    save_dir = os.path.join(DATA_DIR, name)
    ensure_dir(save_dir)
    count = len([f for f in os.listdir(save_dir) if f.lower().endswith(".jpg")])
    out_path = os.path.join(save_dir, f"{count:03d}.jpg")
    cv2.imwrite(out_path, face_roi)

    Student.objects.get_or_create(name=name)
    AuditLog.objects.create(actor="admin", action="enroll", detail=f"Enrolled {name}")
    saved_count = count + 1
    return JsonResponse(
        {
            "ok": True,
            "saved": out_path,
            "count": saved_count,
            "required": REQUIRED_ENROLL_IMAGES,
            "ready_to_train": saved_count >= REQUIRED_ENROLL_IMAGES,
        }
    )


@csrf_exempt
def api_train(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)

    # Reuse train.py logic inline
    faces = []
    labels = []
    id_to_name = {}

    current_id = 0
    for name in sorted(os.listdir(DATA_DIR)):
        person_dir = os.path.join(DATA_DIR, name)
        if not os.path.isdir(person_dir):
            continue

        id_to_name[str(current_id)] = name
        for file_name in os.listdir(person_dir):
            if not file_name.lower().endswith(".jpg"):
                continue
            img_path = os.path.join(person_dir, file_name)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if img is None:
                continue
            faces.append(_normalize_face(img))
            labels.append(current_id)

        current_id += 1

    if not faces:
        return JsonResponse({"ok": False, "error": "No training images found"})

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.train(faces, np.array(labels))

    ensure_dir(MODEL_DIR)
    recognizer.save(MODEL_PATH)
    with open(LABELS_PATH, "w", encoding="utf-8") as f:
        import json

        json.dump(id_to_name, f, indent=2)

    return JsonResponse({"ok": True})


@csrf_exempt
def api_recognize(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    client_ip = request.META.get("REMOTE_ADDR", "unknown")
    if _rate_limited(f"recognize:{client_ip}", limit=180, window_sec=60):
        return JsonResponse({"ok": False, "error": "Too many requests. Please slow down."}, status=429)

    data_url = request.POST.get("frame", "")
    if not data_url:
        return JsonResponse({"ok": False, "error": "Frame required"}, status=400)

    if not (os.path.exists(MODEL_PATH) and os.path.exists(LABELS_PATH)):
        return JsonResponse({"ok": False, "error": "Model not trained"}, status=400)

    labels = load_labels(LABELS_PATH)
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read(MODEL_PATH)

    img = _decode_image(data_url)
    if img is None:
        return JsonResponse({"ok": False, "error": "Invalid image"}, status=400)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(
        os.path.join(cv2.data.haarcascades, "haarcascade_frontalface_default.xml")
    )
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.15, minNeighbors=5, minSize=(60, 60))
    if len(faces) == 0:
        enhanced = cv2.equalizeHist(gray)
        faces = face_cascade.detectMultiScale(enhanced, scaleFactor=1.1, minNeighbors=4, minSize=(45, 45))
    if len(faces) == 0:
        _reset_hits_except(None)
        if ALERTS_ON_UNKNOWN:
            _log_unknown_face(client_ip, "No face match")
        return JsonResponse(
            {"ok": True, "name": "Unknown", "confidence": 999, "marked": False, "message": "Unknown person, please register your face."}
        )
    if len(faces) > 1:
        _reset_hits_except(None)
        return JsonResponse(
            {
                "ok": True,
                "name": "Unknown",
                "confidence": 999,
                "marked": False,
                "message": "Multiple faces detected. Please show only one face.",
            }
        )

    (x, y, w, h) = max(faces, key=lambda f: f[2] * f[3])
    face_roi = gray[y : y + h, x : x + w]

    bright = brightness_level(face_roi)
    if bright < BRIGHTNESS_MIN:
        face_roi = enhance_low_light(face_roi)
        bright = brightness_level(face_roi)
    if bright < BRIGHTNESS_MIN or bright > BRIGHTNESS_MAX:
        _reset_hits_except(None)
        return JsonResponse(
            {
                "ok": True,
                "name": "Unknown",
                "confidence": 999,
                "marked": False,
                "message": "Lighting is not clear. Please face light and try again.",
            }
        )

    # Try multiple preprocessing variants and pick best confidence.
    candidates = []
    base_norm = _normalize_face(face_roi)
    candidates.append(base_norm)
    candidates.append(cv2.equalizeHist(base_norm))
    candidates.append(cv2.GaussianBlur(base_norm, (3, 3), 0))
    if is_blurry(face_roi, BLUR_THRESHOLD)[0]:
        sharpened = cv2.addWeighted(base_norm, 1.5, cv2.GaussianBlur(base_norm, (0, 0), 2), -0.5, 0)
        candidates.append(sharpened)

    best_label = None
    best_confidence = 999.0
    for c in candidates:
        try:
            lbl, conf = recognizer.predict(c)
            if conf < best_confidence:
                best_label = lbl
                best_confidence = conf
        except Exception:
            continue
    if best_label is None:
        _reset_hits_except(None)
        return JsonResponse({"ok": False, "error": "Recognition failed"}, status=500)

    label_id, confidence = best_label, best_confidence
    name = labels.get(str(label_id), "Unknown")
    accept_max = max(min(CONFIDENCE_THRESHOLD, CONFIDENCE_STRICT, RECOGNITION_ACCEPT_MAX), 72)

    if confidence > accept_max:
        name = "Unknown"
        _reset_hits_except(None)

    if name != "Unknown":
        _reset_hits_except(name)
        _last_hits[name] = _last_hits.get(name, 0) + 1
        if _last_hits[name] >= REQUIRED_HITS:
            student = Student.objects.filter(name=name).first()
            if student:
                attendance_type = request.POST.get("attendance_type", "regular")
                makeup_session_id = request.POST.get("makeup_session_id", "").strip()
                if attendance_type == "makeup":
                    if not makeup_session_id:
                        return JsonResponse(
                            {"ok": False, "error": "Make-up session is required for make-up attendance"},
                            status=400,
                        )
                    session = MakeupClassSession.objects.filter(id=makeup_session_id).first()
                    if not session:
                        return JsonResponse({"ok": False, "error": "Invalid make-up session"}, status=400)
                    already_makeup = MakeupAttendance.objects.filter(session=session, student=student).exists()
                    if not already_makeup:
                        MakeupAttendance.objects.create(session=session, student=student, mode="face")
                        AuditLog.objects.create(
                            actor="system",
                            action="makeup_mark",
                            detail=f"{name} {session.remedial_code}",
                        )
                    return JsonResponse(
                        {
                            "ok": True,
                            "name": name,
                            "confidence": int(confidence),
                            "marked": True,
                            "already_marked": already_makeup,
                            "attendance_type": "makeup",
                            "session_code": session.remedial_code,
                        }
                    )

                period = request.POST.get("period", "")
                timetable_entry_id = request.POST.get("timetable_entry_id", "").strip()
                attendance_status = "present"
                if timetable_entry_id:
                    entry = (
                        TimetableEntry.objects.select_related("period", "offering__subject")
                        .filter(id=timetable_entry_id)
                        .first()
                    )
                    if entry:
                        course_code = entry.offering.course.code or entry.offering.course.name
                        subject_name = entry.offering.subject.name
                        slot_label = f"{entry.period.start_time.hour}-{entry.period.end_time.hour}"
                        period = f"{slot_label} - {course_code} - {subject_name}"
                        start_dt = datetime.combine(datetime.now().date(), entry.period.start_time)
                        if datetime.now() > (start_dt + timedelta(minutes=10)):
                            attendance_status = "late"
                if not period:
                    return JsonResponse(
                        {
                            "ok": False,
                            "error": "No timetable configured by admin for today. Select Make-Up Class if applicable.",
                        },
                        status=400,
                    )
                date_val = datetime.now().date()
                existing_att = Attendance.objects.filter(
                    student=student, date=date_val, period_name=period
                ).first()
                already = bool(existing_att)
                if not already:
                    kiosk_id = request.POST.get("kiosk_id", "") or "Webcam"
                    device, _ = Device.objects.get_or_create(name=kiosk_id)
                    device.last_seen_at = datetime.now()
                    device.status = "online"
                    device.save()
                    Attendance.objects.create(
                        student=student,
                        date=date_val,
                        period_name=period,
                        status=attendance_status,
                        device=device,
                    )
                    mark_attendance(name, ATTENDANCE_CSV)
                    AuditLog.objects.create(actor="system", action="mark", detail=f"{name} {period}")
                    if ALERTS_ON_MARKED:
                        _send_notification(
                            subject="Attendance Marked",
                            body=f"{name} marked for {period} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        )
                    if confidence >= max(accept_max - 5, 0):
                        AuditLog.objects.create(
                            actor="system",
                            action="low_conf_mark",
                            detail=f"{name} | {period} | confidence={int(confidence)}",
                        )
                return JsonResponse(
                    {
                        "ok": True,
                        "name": name,
                        "confidence": int(confidence),
                        "marked": True,
                        "already_marked": already,
                        "mark_status": existing_att.status if existing_att else attendance_status,
                    }
                )
            _last_hits[name] = 0
            _last_marked[name] = time_now()

    if name == "Unknown":
        if ALERTS_ON_UNKNOWN:
            _log_unknown_face(client_ip, "Unknown prediction")
            _send_notification(
                subject="Unknown Face Attempt",
                body=f"Unknown face detected on kiosk webcam at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            )
        return JsonResponse(
            {"ok": True, "name": "Unknown", "confidence": int(confidence), "marked": False, "already_marked": False, "message": "Unknown person, please register your face."}
        )
    return JsonResponse(
        {
            "ok": True,
            "name": name,
            "confidence": int(confidence),
            "marked": False,
            "already_marked": False,
            "message": "Face detected. Hold steady for attendance.",
        }
    )


def api_status(request):
    return JsonResponse({"ok": True, "status": _last_status})


def time_now():
    return datetime.now().timestamp()


def api_recent(request):
    view_type = request.GET.get("type", "all").strip().lower()
    date_str = request.GET.get("date", "").strip()
    if date_str:
        try:
            date_val = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse({"ok": False, "error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
    else:
        date_val = datetime.now().date()
    data = []

    if view_type in {"all", "regular"}:
        regular_records = (
            Attendance.objects.select_related("student", "device")
            .filter(date=date_val)
            .order_by("-marked_at")[:50]
        )
        for r in regular_records:
            data.append(
                {
                    "name": r.student.name,
                    "period": r.period_name,
                    "timestamp": r.marked_at.strftime("%Y-%m-%d %H:%M:%S"),
                    "camera": r.device.name if r.device else "Webcam",
                    "_sort": r.marked_at.timestamp(),
                }
            )

    if view_type in {"all", "makeup"}:
        makeup_records = (
            MakeupAttendance.objects.select_related("student", "session")
            .filter(session__date=date_val)
            .order_by("-marked_at")[:50]
        )
        for m in makeup_records:
            session = m.session
            period_label = f"Make-Up | {session.remedial_code}" if session else "Make-Up"
            data.append(
                {
                    "name": m.student.name,
                    "period": period_label,
                    "timestamp": m.marked_at.strftime("%Y-%m-%d %H:%M:%S"),
                    "camera": "Webcam",
                    "_sort": m.marked_at.timestamp(),
                }
            )

    data.sort(key=lambda x: x["_sort"], reverse=True)
    data = data[:50]
    for row in data:
        row.pop("_sort", None)
    return JsonResponse({"ok": True, "records": data})


@csrf_exempt
def api_export(request):
    from openpyxl import Workbook

    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)

    date_str = request.GET.get("date", datetime.now().strftime("%Y-%m-%d"))
    date_val = datetime.strptime(date_str, "%Y-%m-%d").date()
    records = (
        Attendance.objects.select_related("student", "device")
        .filter(date=date_val)
        .order_by("marked_at")
    )
    if not records:
        return JsonResponse({"ok": False, "error": "No attendance for date"})

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    logo_path = os.path.join("webapp", "static", "lpu-logo.jfif")
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image

            img = Image(logo_path)
            img.height = 60
            img.width = 60
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws["B1"] = "Lovely Professional University"
    ws["B2"] = f"Attendance Report - {date_str}"
    ws.append([])
    ws.append(["Name", "Period", "Timestamp", "Camera"])
    for r in records:
        ws.append(
            [
                r.student.name,
                r.period_name,
                r.marked_at.strftime("%Y-%m-%d %H:%M:%S"),
                r.device.name if r.device else "Webcam",
            ]
        )

    ws.append([])
    ws.append(["Faculty Signature:", "___________________"])
    ws.append(["HOD Signature:", "___________________"])

    out_path = os.path.join("reports", f"attendance_{date_str}.xlsx")
    ensure_dir(os.path.dirname(out_path))
    wb.save(out_path)

    return JsonResponse({"ok": True, "path": out_path})


@csrf_exempt
def api_export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader

    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)

    date_str = request.GET.get("date", datetime.now().strftime("%Y-%m-%d"))
    date_val = datetime.strptime(date_str, "%Y-%m-%d").date()
    records = (
        Attendance.objects.select_related("student", "device")
        .filter(date=date_val)
        .order_by("marked_at")
    )
    if not records:
        return JsonResponse({"ok": False, "error": "No attendance for date"})

    out_path = os.path.join("reports", f"attendance_{date_str}.pdf")
    ensure_dir(os.path.dirname(out_path))
    c = canvas.Canvas(out_path, pagesize=A4)
    width, height = A4
    y = height - 50

    logo_path = os.path.join("webapp", "static", "lpu-logo.jfif")
    if os.path.exists(logo_path):
        try:
            c.drawImage(ImageReader(logo_path), 40, y - 40, width=50, height=50, mask="auto")
        except Exception:
            pass

    c.setFont("Helvetica-Bold", 14)
    c.drawString(100, y, "Lovely Professional University")
    y -= 18
    c.setFont("Helvetica-Bold", 14)
    c.drawString(100, y, f"Attendance Report - {date_str}")
    y -= 30
    c.setFont("Helvetica", 10)
    c.drawString(40, y, "Name")
    c.drawString(220, y, "Period")
    c.drawString(320, y, "Timestamp")
    c.drawString(480, y, "Camera")
    y -= 15

    for r in records:
        if y < 50:
            c.showPage()
            y = height - 50
        c.drawString(40, y, r.student.name)
        c.drawString(220, y, r.period_name)
        c.drawString(320, y, r.marked_at.strftime("%Y-%m-%d %H:%M:%S"))
        c.drawString(480, y, r.device.name if r.device else "Webcam")
        y -= 14

    y -= 30
    c.drawString(40, y, "Faculty Signature: ___________________")
    y -= 18
    c.drawString(40, y, "HOD Signature: ___________________")
    c.save()
    return JsonResponse({"ok": True, "path": out_path})


@csrf_exempt
def api_reset_today(request):
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    date_val = datetime.now().date()
    Attendance.objects.filter(date=date_val).delete()
    AuditLog.objects.create(actor="admin", action="reset", detail=f"Cleared {date_val.isoformat()}")
    return JsonResponse({"ok": True, "cleared": date_val.isoformat()})


def api_dashboard(request):
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    today = datetime.now().date()
    total_students = Student.objects.count()
    today_marked = Attendance.objects.filter(date=today).count()
    defaulters = Student.objects.exclude(
        id__in=Attendance.objects.filter(date=today).values_list("student_id", flat=True)
    )[:20]
    attendance_rate = round((today_marked / total_students) * 100, 2) if total_students else 0

    weekly_trend = []
    for offset in range(6, -1, -1):
        day = today - timedelta(days=offset)
        day_marked = Attendance.objects.filter(date=day).count()
        day_rate = round((day_marked / total_students) * 100, 2) if total_students else 0
        weekly_trend.append(
            {
                "date": day.strftime("%Y-%m-%d"),
                "marked": day_marked,
                "rate": day_rate,
            }
        )

    month_start = today.replace(day=1)
    monthly_counts = (
        Attendance.objects.filter(date__gte=month_start, date__lte=today)
        .values("student__name")
        .annotate(count=models.Count("id"))
        .order_by("count", "student__name")[:10]
    )

    period_wise = (
        Attendance.objects.filter(date=today)
        .values("period_name")
        .annotate(count=models.Count("id"))
        .order_by("-count", "period_name")[:8]
    )

    data = {
        "total_students": total_students,
        "today_marked": today_marked,
        "attendance_rate": attendance_rate,
        "today_defaulters": [s.name for s in defaulters],
        "weekly_trend": list(weekly_trend),
        "monthly_low_attendance": list(monthly_counts),
        "period_wise_today": list(period_wise),
    }
    return JsonResponse({"ok": True, "data": data})


def api_devices(request):
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    devices = Device.objects.all().order_by("name")
    data = [
        {
            "name": d.name,
            "location": d.location,
            "status": d.status,
            "last_seen_at": d.last_seen_at.strftime("%Y-%m-%d %H:%M:%S") if d.last_seen_at else "",
        }
        for d in devices
    ]
    return JsonResponse({"ok": True, "devices": data})


def api_defaulters(request):
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    date_val = datetime.now().date()
    defaulters = Student.objects.exclude(
        id__in=Attendance.objects.filter(date=date_val).values_list("student_id", flat=True)
    )
    data = [s.name for s in defaulters]
    return JsonResponse({"ok": True, "defaulters": data})


def api_monthly_report(request):
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    today = datetime.now().date()
    month_start = today.replace(day=1)
    month_attendance = Attendance.objects.filter(date__gte=month_start, date__lte=today)
    total = month_attendance.count()
    per_student = (
        month_attendance.values("student__name")
        .order_by("student__name")
        .annotate(count=models.Count("id"))
    )
    data = {
        "total_records": total,
        "per_student": list(per_student),
        "month": month_start.strftime("%Y-%m"),
    }
    return JsonResponse({"ok": True, "data": data})


def api_makeup_sessions(request):
    today = datetime.now().date()
    sessions = MakeupClassSession.objects.select_related("faculty", "course", "class_section").filter(
        date__gte=today
    ).order_by("date", "start_time")[:50]
    data = [
        {
            "id": s.id,
            "date": s.date.strftime("%Y-%m-%d"),
            "start_time": s.start_time.strftime("%H:%M"),
            "end_time": s.end_time.strftime("%H:%M"),
            "faculty": s.faculty.name if s.faculty else "",
            "course": s.course.name if s.course else "",
            "section": s.class_section.name if s.class_section else "",
            "remedial_code": s.remedial_code,
        }
        for s in sessions
    ]
    return JsonResponse({"ok": True, "sessions": data})


@csrf_exempt
def api_makeup_create(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    if not _has_faculty_access(request):
        return JsonResponse({"ok": False, "error": "Lecturer login required"}, status=403)
    faculty_name = request.POST.get("faculty_name", "").strip()
    course_name = request.POST.get("course_name", "").strip()
    section_name = request.POST.get("section_name", "").strip()
    date_str = request.POST.get("date", "").strip()
    start_time = request.POST.get("start_time", "").strip()
    end_time = request.POST.get("end_time", "").strip()
    notes = request.POST.get("notes", "").strip()

    if not date_str or not start_time or not end_time:
        return JsonResponse({"ok": False, "error": "Date and time are required"}, status=400)

    try:
        date_val = datetime.strptime(date_str, "%Y-%m-%d").date()
        start_val = datetime.strptime(start_time, "%H:%M").time()
        end_val = datetime.strptime(end_time, "%H:%M").time()
    except ValueError:
        return JsonResponse({"ok": False, "error": "Invalid date/time format"}, status=400)

    faculty = Faculty.objects.filter(name=faculty_name).first() if faculty_name else None
    if faculty_name and not faculty:
        faculty = Faculty.objects.create(name=faculty_name)
    course = Course.objects.filter(name=course_name).first() if course_name else None
    if course_name and not course:
        course = Course.objects.create(name=course_name, code=course_name[:10].upper())
    section = ClassSection.objects.filter(name=section_name).first() if section_name else None
    if section_name and not section:
        section = ClassSection.objects.create(name=section_name)

    session = MakeupClassSession.objects.create(
        faculty=faculty,
        course=course,
        class_section=section,
        date=date_val,
        start_time=start_val,
        end_time=end_val,
        notes=notes,
    )
    return JsonResponse({"ok": True, "session_id": session.id, "remedial_code": session.remedial_code})


@csrf_exempt
def api_place_food_order(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    student_name = request.POST.get("student_name", "").strip()
    slot_id = request.POST.get("slot_id", "").strip()
    order_mode = request.POST.get("order_mode", "dining").strip().lower()
    delivery_location = request.POST.get("delivery_location", "").strip()
    payment_method = "qr"
    items_raw = request.POST.get("items", "[]")

    if not student_name:
        return JsonResponse({"ok": False, "error": "Student name is required"}, status=400)
    if order_mode not in {"dining", "delivery"}:
        return JsonResponse({"ok": False, "error": "Invalid order mode"}, status=400)
    if order_mode == "delivery" and not delivery_location:
        return JsonResponse({"ok": False, "error": "Delivery location required"}, status=400)

    try:
        items_data = json.loads(items_raw)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Invalid items payload"}, status=400)
    if not items_data:
        return JsonResponse({"ok": False, "error": "Select at least one food item"}, status=400)

    slot = BreakTimeSlot.objects.filter(id=slot_id).first() if slot_id else None
    student, _ = Student.objects.get_or_create(name=student_name)

    is_qr_payment = True
    effective_payment_status = "pending"

    order = FoodOrder.objects.create(
        student=student,
        slot=slot,
        order_mode=order_mode,
        delivery_location=delivery_location,
        status="pending" if is_qr_payment else "confirmed",
    )
    total_amount = 0
    total_prep_mins = 0
    for item_data in items_data:
        food_id = item_data.get("food_item_id")
        quantity = int(item_data.get("quantity", 1))
        if quantity <= 0:
            continue
        food_item = FoodItem.objects.filter(id=food_id, is_available=True).first()
        if not food_item:
            continue
        FoodOrderItem.objects.create(
            order=order,
            food_item=food_item,
            quantity=quantity,
            unit_price=food_item.price,
        )
        total_amount += float(food_item.price) * quantity
        total_prep_mins += food_item.prep_time_mins * quantity

    if total_amount <= 0:
        order.delete()
        return JsonResponse({"ok": False, "error": "No valid food items selected"}, status=400)

    order.total_amount = total_amount
    order.save()
    merchant_txn_id = f"PP-{order.id}-{int(datetime.now().timestamp())}"
    payment = Payment.objects.create(
        order=order,
        method="qr",
        status="pending",
        amount=total_amount,
        transaction_id=merchant_txn_id,
        paid_at=None,
    )
    slot_queue = FoodOrder.objects.filter(
        slot=slot,
        status__in=["pending", "confirmed", "preparing"],
    ).exclude(id=order.id)
    queue_count = slot_queue.count()
    queue_wait = queue_count * 4
    predicted_wait_mins = int(total_prep_mins + queue_wait)
    eta_at = timezone.now() + timedelta(minutes=predicted_wait_mins)
    response = {
        "ok": True,
        "order_id": order.id,
        "payment_id": payment.id,
        "payment_status": payment.status,
        "total_amount": order.total_amount,
        "predicted_wait_mins": predicted_wait_mins,
        "predicted_eta": eta_at.strftime("%Y-%m-%d %H:%M"),
    }
    return JsonResponse(response)


@csrf_exempt
def api_food_update_status(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    order_id = request.POST.get("order_id", "").strip()
    new_status = request.POST.get("status", "").strip()
    allowed = {"pending", "confirmed", "preparing", "ready", "collected", "cancelled"}
    if not order_id or new_status not in allowed:
        return JsonResponse({"ok": False, "error": "order_id and valid status required"}, status=400)
    order = FoodOrder.objects.filter(id=order_id).first()
    if not order:
        return JsonResponse({"ok": False, "error": "Order not found"}, status=404)
    order.status = new_status
    order.save()
    AuditLog.objects.create(actor="admin", action="food_status_update", detail=f"Order {order.id} -> {new_status}")
    return JsonResponse({"ok": True, "order_id": order.id, "status": order.status})


@csrf_exempt
def api_qr_payment_received(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    order_id = request.POST.get("order_id", "").strip()
    payment_id = request.POST.get("payment_id", "").strip()
    scanner_ref = request.POST.get("scanner_ref", "").strip()
    received_amount_raw = request.POST.get("received_amount", "").strip()
    scanner_token = request.POST.get("scanner_token", "").strip()
    if not QR_SCANNER_TOKEN:
        return JsonResponse({"ok": False, "error": "QR scanner integration is not configured"}, status=503)
    if scanner_token != QR_SCANNER_TOKEN:
        return JsonResponse({"ok": False, "error": "Invalid scanner token"}, status=403)
    if not order_id and not payment_id:
        return JsonResponse({"ok": False, "error": "order_id or payment_id required"}, status=400)

    payment_qs = Payment.objects.select_related("order")
    payment = None
    if payment_id:
        payment = payment_qs.filter(id=payment_id).first()
    elif order_id:
        payment = payment_qs.filter(order_id=order_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)
    if payment.method != "qr":
        return JsonResponse({"ok": False, "error": "Payment is not QR mode"}, status=400)
    if payment.status == "paid":
        return JsonResponse(
            {
                "ok": True,
                "order_id": payment.order_id,
                "payment_id": payment.id,
                "payment_status": payment.status,
                "transaction_id": payment.transaction_id,
            }
        )
    if not received_amount_raw:
        return JsonResponse({"ok": False, "error": "received_amount required"}, status=400)
    try:
        received_amount = float(received_amount_raw)
    except ValueError:
        return JsonResponse({"ok": False, "error": "Invalid received_amount"}, status=400)
    expected_amount = float(payment.amount)
    if abs(received_amount - expected_amount) > 0.009:
        return JsonResponse(
            {
                "ok": False,
                "error": "Amount mismatch",
                "expected_amount": expected_amount,
                "received_amount": received_amount,
            },
            status=400,
        )

    payment.status = "paid"
    payment.paid_at = timezone.now()
    if scanner_ref:
        payment.transaction_id = scanner_ref
    payment.save()

    order = payment.order
    if order:
        order.status = "confirmed"
        order.save()
    AuditLog.objects.create(
        actor="scanner",
        action="qr_payment_received",
        detail=f"Order {order.id if order else '-'} Payment {payment.id}",
    )
    return JsonResponse(
        {
            "ok": True,
            "order_id": order.id if order else None,
            "payment_id": payment.id,
            "payment_status": payment.status,
            "transaction_id": payment.transaction_id,
        }
    )


def api_payment_status(request):
    order_id = request.GET.get("order_id", "").strip()
    payment_id = request.GET.get("payment_id", "").strip()
    if not order_id and not payment_id:
        return JsonResponse({"ok": False, "error": "order_id or payment_id required"}, status=400)
    payment_qs = Payment.objects.select_related("order")
    payment = payment_qs.filter(id=payment_id).first() if payment_id else payment_qs.filter(order_id=order_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)
    return JsonResponse(
        {
            "ok": True,
            "payment_id": payment.id,
            "order_id": payment.order_id,
            "payment_status": payment.status,
            "method": payment.method,
            "transaction_id": payment.transaction_id,
        }
    )


@csrf_exempt
def api_submit_upi_proof(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    order_id = request.POST.get("order_id", "").strip()
    payment_id = request.POST.get("payment_id", "").strip()
    upi_txn_id = request.POST.get("upi_txn_id", "").strip()
    received_amount_raw = request.POST.get("received_amount", "").strip()

    if not order_id and not payment_id:
        return JsonResponse({"ok": False, "error": "order_id or payment_id required"}, status=400)
    if not upi_txn_id:
        return JsonResponse({"ok": False, "error": "UPI transaction ID required"}, status=400)
    if not re.match(r"^[A-Za-z0-9_-]{8,60}$", upi_txn_id):
        return JsonResponse({"ok": False, "error": "Invalid UPI transaction ID format"}, status=400)
    if not received_amount_raw:
        return JsonResponse({"ok": False, "error": "received_amount required"}, status=400)
    try:
        received_amount = float(received_amount_raw)
    except ValueError:
        return JsonResponse({"ok": False, "error": "Invalid received_amount"}, status=400)

    payment_qs = Payment.objects.select_related("order")
    payment = payment_qs.filter(id=payment_id).first() if payment_id else payment_qs.filter(order_id=order_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)
    if payment.status == "paid":
        return JsonResponse({"ok": False, "error": "Payment already marked as paid"}, status=400)
    if Payment.objects.filter(transaction_id=upi_txn_id).exclude(id=payment.id).exists():
        return JsonResponse({"ok": False, "error": "Duplicate transaction ID"}, status=400)
    expected_amount = float(payment.amount)
    if abs(received_amount - expected_amount) > 0.009:
        return JsonResponse(
            {
                "ok": False,
                "error": "Amount mismatch",
                "expected_amount": expected_amount,
                "received_amount": received_amount,
            },
            status=400,
        )

    payment.transaction_id = upi_txn_id
    payment.status = "pending"
    payment.save()
    AuditLog.objects.create(
        actor="student",
        action="upi_proof_submitted",
        detail=f"Order {payment.order_id} Txn {upi_txn_id}",
    )
    return JsonResponse({"ok": True, "payment_id": payment.id, "payment_status": payment.status})


@csrf_exempt
def api_admin_payment_review(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)

    payment_id = request.POST.get("payment_id", "").strip()
    decision = request.POST.get("decision", "").strip().lower()
    if not payment_id or decision not in {"received", "not_received"}:
        return JsonResponse({"ok": False, "error": "payment_id and valid decision required"}, status=400)

    payment = Payment.objects.select_related("order").filter(id=payment_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)

    if decision == "received":
        payment.status = "paid"
        payment.paid_at = timezone.now()
        payment.save()
        if payment.order:
            payment.order.status = "confirmed"
            payment.order.save()
        AuditLog.objects.create(actor="admin", action="payment_received", detail=f"Payment {payment.id}")
        _send_notification(
            subject="Payment Verified",
            body=f"Payment #{payment.id} was marked as received by admin.",
        )
    else:
        payment.status = "failed"
        payment.save()
        if payment.order:
            payment.order.status = "cancelled"
            payment.order.save()
        AuditLog.objects.create(actor="admin", action="payment_not_received", detail=f"Payment {payment.id}")
        _send_notification(
            subject="Payment Failed",
            body=f"Payment #{payment.id} was marked as not received by admin.",
        )

    return JsonResponse({"ok": True, "payment_id": payment.id, "payment_status": payment.status})


@csrf_exempt
def api_payment_dispute(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    payment_id = request.POST.get("payment_id", "").strip()
    note = request.POST.get("note", "").strip()
    if not payment_id or not note:
        return JsonResponse({"ok": False, "error": "payment_id and note required"}, status=400)
    payment = Payment.objects.select_related("order").filter(id=payment_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)
    AuditLog.objects.create(
        actor="admin",
        action="payment_dispute",
        detail=f"Payment {payment.id} Order {payment.order_id} :: {note[:250]}",
    )
    return JsonResponse({"ok": True, "payment_id": payment.id})


@csrf_exempt
def api_payment_refund(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    payment_id = request.POST.get("payment_id", "").strip()
    note = request.POST.get("note", "").strip()
    if not payment_id:
        return JsonResponse({"ok": False, "error": "payment_id required"}, status=400)
    payment = Payment.objects.select_related("order").filter(id=payment_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)
    if payment.status != "paid":
        return JsonResponse({"ok": False, "error": "Only paid payments can be refunded"}, status=400)
    payment.status = "refunded"
    payment.save()
    if payment.order:
        payment.order.status = "cancelled"
        payment.order.save()
    AuditLog.objects.create(
        actor="admin",
        action="payment_refund",
        detail=f"Payment {payment.id} refunded. {note[:200]}",
    )
    return JsonResponse({"ok": True, "payment_id": payment.id, "payment_status": payment.status})


@csrf_exempt
def api_payment_upload_proof(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    payment_id = request.POST.get("payment_id", "").strip()
    proof = request.FILES.get("proof")
    if not payment_id or not proof:
        return JsonResponse({"ok": False, "error": "payment_id and proof file required"}, status=400)
    payment = Payment.objects.filter(id=payment_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)
    if proof.size > (2 * 1024 * 1024):
        return JsonResponse({"ok": False, "error": "Proof file too large. Max 2MB"}, status=400)
    ext = os.path.splitext(proof.name)[1].lower()
    if ext not in {".png", ".jpg", ".jpeg", ".pdf"}:
        return JsonResponse({"ok": False, "error": "Allowed formats: png, jpg, jpeg, pdf"}, status=400)
    ensure_dir(os.path.join(DATA_DIR, "payment_proofs"))
    safe_name = f"payment_{payment.id}_{int(datetime.now().timestamp())}{ext}"
    target_path = os.path.join(DATA_DIR, "payment_proofs", safe_name)
    with open(target_path, "wb") as f:
        for chunk in proof.chunks():
            f.write(chunk)
    AuditLog.objects.create(
        actor="student",
        action="payment_proof_uploaded",
        detail=f"Payment {payment.id} proof {safe_name}",
    )
    return JsonResponse({"ok": True, "payment_id": payment.id, "proof_file": safe_name})


def api_schedule_suggestions(request):
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    day = request.GET.get("day", "").strip()
    faculty_name = request.GET.get("faculty_name", "").strip()
    room = request.GET.get("room", "").strip()
    section_name = request.GET.get("section_name", "").strip()
    if not day:
        return JsonResponse({"ok": False, "error": "day required"}, status=400)
    faculty = Faculty.objects.filter(name__iexact=faculty_name).first() if faculty_name else None
    section = ClassSection.objects.filter(name__iexact=section_name).first() if section_name else None
    suggestions = _suggest_next_slots(day=day, period=None, room=room, faculty=faculty, section=section, limit=6)
    return JsonResponse({"ok": True, "day": day, "suggested_slots": suggestions})


def api_analytics_summary(request):
    if not _has_admin_access(request):
        return JsonResponse({"ok": False, "error": "Admin required"}, status=403)
    today = datetime.now().date()
    week_start = today - timedelta(days=6)
    att_qs = Attendance.objects.filter(date__gte=week_start, date__lte=today)
    attendance_by_day = (
        att_qs.values("date").annotate(total=models.Count("id")).order_by("date")
    )
    food_peak = (
        FoodOrder.objects.values("slot__name")
        .annotate(total=models.Count("id"))
        .order_by("-total")[:8]
    )
    payment_split = (
        Payment.objects.values("status")
        .annotate(total=models.Count("id"), amount=models.Sum("amount"))
        .order_by("status")
    )
    return JsonResponse(
        {
            "ok": True,
            "attendance_by_day": list(attendance_by_day),
            "food_peak_slots": list(food_peak),
            "payment_split": list(payment_split),
        }
    )


@csrf_exempt
def api_phonepe_webhook(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    token = request.GET.get("token", "").strip() or request.POST.get("token", "").strip()
    if not PHONEPE_WEBHOOK_TOKEN:
        return JsonResponse({"ok": False, "error": "PHONEPE_WEBHOOK_TOKEN not configured"}, status=503)
    if token != PHONEPE_WEBHOOK_TOKEN:
        return JsonResponse({"ok": False, "error": "Invalid webhook token"}, status=403)

    try:
        body = json.loads(request.body.decode("utf-8") if request.body else "{}")
    except json.JSONDecodeError:
        body = {}
    merchant_txn_id = (
        body.get("merchantTransactionId")
        or body.get("transactionId")
        or request.POST.get("merchantTransactionId", "").strip()
        or request.POST.get("transactionId", "").strip()
    )
    amount_raw = body.get("amount") or request.POST.get("amount", "").strip()
    state = (body.get("state") or body.get("status") or request.POST.get("state", "")).strip().upper()
    provider_ref = body.get("providerReferenceId") or request.POST.get("providerReferenceId", "").strip()

    if not merchant_txn_id:
        return JsonResponse({"ok": False, "error": "merchantTransactionId required"}, status=400)
    payment = Payment.objects.select_related("order").filter(transaction_id=merchant_txn_id).first()
    if not payment:
        return JsonResponse({"ok": False, "error": "Payment not found"}, status=404)
    if payment.method != "qr":
        return JsonResponse({"ok": False, "error": "Payment is not QR mode"}, status=400)

    try:
        received_amount = float(amount_raw) if amount_raw not in (None, "") else float(payment.amount)
    except ValueError:
        return JsonResponse({"ok": False, "error": "Invalid amount"}, status=400)
    expected_amount = float(payment.amount)
    if abs(received_amount - expected_amount) > 0.009:
        return JsonResponse(
            {
                "ok": False,
                "error": "Amount mismatch",
                "expected_amount": expected_amount,
                "received_amount": received_amount,
            },
            status=400,
        )

    if state in {"COMPLETED", "SUCCESS", "PAID"}:
        payment.status = "paid"
        payment.paid_at = timezone.now()
        if provider_ref:
            payment.transaction_id = merchant_txn_id
        payment.save()
        if payment.order:
            payment.order.status = "confirmed"
            payment.order.save()
        AuditLog.objects.create(
            actor="phonepe_webhook",
            action="payment_success",
            detail=f"Order {payment.order_id} Txn {merchant_txn_id}",
        )
        return JsonResponse({"ok": True, "payment_status": "paid", "order_id": payment.order_id})

    if state in {"FAILED", "DECLINED"}:
        payment.status = "failed"
        payment.save()
        if payment.order:
            payment.order.status = "cancelled"
            payment.order.save()
        return JsonResponse({"ok": True, "payment_status": "failed", "order_id": payment.order_id})

    return JsonResponse({"ok": True, "payment_status": payment.status, "order_id": payment.order_id})


@csrf_exempt
def api_lecturer_login(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    username = request.POST.get("username", "").strip()
    password = request.POST.get("password", "").strip()
    user = auth.authenticate(request, username=username, password=password)
    if not user:
        return JsonResponse({"ok": False, "error": "Invalid credentials"}, status=401)
    if not _user_is_faculty(user):
        return JsonResponse({"ok": False, "error": "Faculty role required"}, status=403)
    request.session["lecturer_user_id"] = user.id
    request.session["lecturer_name"] = user.get_username()
    return JsonResponse({"ok": True, "name": user.get_username()})


@csrf_exempt
def api_lecturer_logout(request):
    request.session["lecturer_user_id"] = None
    request.session["lecturer_name"] = ""
    return JsonResponse({"ok": True})


@csrf_exempt
def api_admin_login(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    client_ip = request.META.get("REMOTE_ADDR", "unknown")
    if _rate_limited(f"admin_login:{client_ip}", limit=8, window_sec=60):
        return JsonResponse({"ok": False, "error": "Too many login attempts. Try again in a minute."}, status=429)

    username = request.POST.get("username", "").strip() or ADMIN_USERNAME
    password = request.POST.get("password", "")
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        request.session["is_admin"] = True
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "error": "Invalid password"}, status=401)


@csrf_exempt
def api_admin_logout(request):
    request.session["is_admin"] = False
    request.session["lecturer_user_id"] = None
    request.session["lecturer_name"] = ""
    auth.logout(request)
    return JsonResponse({"ok": True})
